from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from PIL import Image,ImageDraw,ImageFont
import webcolors
import os
import openpyxl
import tkinter as tk
import json


long_colors = []
colors = []

def img_to_grid(img, dim_x, dim_y):
    grid = []
    palette = []
    with Image.open(img) as image:
        w, h = image.size  # get image dimensions
        pixels = image.load()
        row_w = int(w / dim_x)  # define dimensions of "stitch" based on image dimensions
        row_h = int(h / dim_y)

        for x in range(dim_x):
            row = []
            for y in range(dim_y):
                box = (y * row_w, x * row_h, y * row_w + row_w, x * row_h + row_h)  # define box (range of pixels) to crop out of original image
                output = image.crop(box)  # crop
                row.append(get_avg_color(output))

                if get_avg_color(output) not in palette:  # append unique colors to array for color palette
                    palette.append(get_avg_color(output))
                else:
                    continue

            grid.append(row)

    return grid, palette


def grid_to_sheet(img_path, grid, long_colors):
    sheet = f'{img_path.split("/")[-1].split(".")[0]}.xlsx'  # spreadsheet to save pattern to
    wb = Workbook()
    ws = wb.active
    border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

    for x in range(len(grid)):
        for y in range(len(grid[x])):
            ws.column_dimensions[f'{get_column_letter(y+1)}'].width = 2.6  # set column width to more closely match height
            cell = ws[f'{get_column_letter(y+1)}{x+1}']
            cell.fill = PatternFill("solid", fgColor=f'{grid[x][y].strip("#")}')  # fill cell with color from image
            cell.border = border

    wb.save(sheet)

    colors = list(set(long_colors))
    font = ImageFont.load_default().font
    palette = Image.new(mode="RGB", size=(1000,120))  # create blank image for color palette

    for i in range(len(colors)):
        newt = Image.new(mode="RGB", size=(1000//len(colors),100), color=colors[i])
        draw = ImageDraw.Draw(newt)
        draw.text((5,5), str(colors[i]), fill="white", font=font)  # write text for hex code
        palette.paste(newt, (i*1000//len(colors),10))  # paste color onto image

    palette.save("palette.png")


def get_avg_color(img):  # scan the image and return the most common color in hex
    long_colors = []
    w, h = img.size
    pixels = img.load()

    for y in range(h):
        for x in range(w):
            try:
                #print(pixels[x,y])
                r, g, b, a = pixels[x,y]
                if a == 0:
                    long_colors.append(webcolors.rgb_to_hex((255,255,255))) #white for blank pixels
                else:
                    long_colors.append(webcolors.rgb_to_hex((r, g, b)))
            except TypeError:
                print(pixels[x,y])

    most_common = max(set(long_colors), key = long_colors.count)

    return most_common


def get_palette(img):
    with Image.open(img) as image:
        w, h = image.size
        pixels = image.load()
        for y in range(h):
            for x in range(w):
                try:
                    #print(pixels[x,y])
                    r, g, b, a = pixels[x,y]
                    long_colors.append(webcolors.rgb_to_hex((r, g, b)))
                except TypeError:
                    print(pixels[x,y])

    colors = list(set(long_colors))
    font = ImageFont.load_default().font
    palette = Image.new(mode="RGB", size=(1000,120))

    for i in range(len(colors)):
        newt = Image.new(mode="RGB", size=(1000//len(colors),100), color=colors[i])
        draw = ImageDraw.Draw(newt)
        draw.text((5,5), str(colors[i]), fill="white", font=font)
        palette.paste(newt, (i*1000//len(colors),10))

    palette.save("palette.png")


def nearest_color(rgb):
    if rgb in colors:
        return colors[x]
    elif rgb not in colors:
        for x in colors:
            if colors[x] - rgb <= 10:
                return colors[x]


def submit():
    img_path = p_input.get().strip()
    dim_x, dim_y = dim_input.get().strip().split("x")
    grid, palette = img_to_grid(img_path, int(dim_x), int(dim_y))
    grid_to_sheet(img_path, grid, palette)


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Stitch")

    p_label = tk.Label(root, text="Path to image: ").grid(row=0,column=0)
    p_input = tk.Entry(root)
    p_input.grid(row=0,column=1)

    dim_label = tk.Label(root, text="Stitch Dimensions: ").grid(row=1,column=0)
    dim_input = tk.Entry(root)
    dim_input.grid(row=1,column=1)
    dim_label2 = tk.Label(root, text="(ex: 20x20)").grid(row=2,column=1)

    submit_button = tk.Button(root, text="Submit", command=submit).grid(row=3,column=1)

    root.update_idletasks()
    root.mainloop()
